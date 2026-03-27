import json
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ── Load JSON ────────────────────────────────────────────────────────────────
JSON_PATH  = r"C:\Users\ChaitanyaMalle\CF AI\eval\questions.json"
EXCEL_PATH = r"C:\Users\ChaitanyaMalle\CF AI\eval\CloudFuze_Manage_AI_Eval_Questions.xlsx"

with open(JSON_PATH, "r", encoding="utf-8") as f:
    questions = json.load(f)

print(f"Loaded {len(questions)} questions")

# ── Helpers ──────────────────────────────────────────────────────────────────
def make_fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)

def make_font(bold=False, color="000000", name="Arial", size=11):
    return Font(bold=bold, color=color, name=name, size=size)

def make_border():
    thin = Side(style="thin", color="D1D5DB")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

# ── Colour maps ──────────────────────────────────────────────────────────────
DOMAIN_FILL = {
    "licence":      "DBEAFE",
    "spend":        "DCFCE7",
    "shadow_it":    "FEE2E2",
    "contracts":    "FEF9C3",
    "users":        "F3E8FF",
    "compliance":   "FFEDD5",
    "multi":        "CCFBF1",
    "out_of_scope": "F3F4F6",
}
DIFF_COLOR = {
    "easy":   "16A34A",
    "medium": "D97706",
    "hard":   "DC2626",
}
SCOPE_FILL = {
    "TRUE":  "DCFCE7",
    "FALSE": "FEE2E2",
}

HEADER_FILL = make_fill("1E3A5F")
HEADER_FONT = make_font(bold=True, color="FFFFFF", name="Arial", size=11)
ALT_ROW_FILL = make_fill("F9FAFB")

# ── Workbook ─────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ============================================================
# SHEET 1 — All Questions
# ============================================================
ws1 = wb.active
ws1.title = "All Questions"

HEADERS = [
    "ID", "Question", "Domain", "Intent", "Expected Widget",
    "Expected Tools", "In Scope", "Difficulty", "Source", "Notes",
    "Test Result", "Pass/Fail", "Score (1-5)", "Comments"
]

COL_WIDTHS = {
    1: 10, 2: 60, 3: 15, 4: 15, 5: 18,
    6: 40, 7: 10, 8: 12, 9: 10, 10: 40,
    11: 15, 12: 12, 13: 12, 14: 30
}

# Header row
for col_idx, header in enumerate(HEADERS, 1):
    cell = ws1.cell(row=1, column=col_idx, value=header)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = make_border()

ws1.row_dimensions[1].height = 30

# Data rows
for row_idx, q in enumerate(questions, 2):
    tools_str = ", ".join(q.get("expected_tools", []))
    in_scope_str = "TRUE" if q.get("in_scope", True) else "FALSE"
    is_even = (row_idx % 2 == 0)

    row_data = [
        q.get("id", ""),
        q.get("question", ""),
        q.get("domain", ""),
        q.get("intent", ""),
        q.get("expected_widget", ""),
        tools_str,
        in_scope_str,
        q.get("difficulty", ""),
        q.get("source", "manual"),
        q.get("notes", ""),
        "", "", "", ""          # K–N empty
    ]

    for col_idx, value in enumerate(row_data, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=value)
        cell.font = make_font(name="Arial", size=10)
        cell.border = make_border()

        # Default alternate row shading (applied first, overridden below)
        if is_even:
            cell.fill = ALT_ROW_FILL

        # Column-specific formatting
        if col_idx == 2:  # Question — wrap
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        elif col_idx == 10:  # Notes — wrap
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        else:
            cell.alignment = Alignment(vertical="center")

        # Domain colour (col C = 3)
        if col_idx == 3:
            domain = value.lower()
            if domain in DOMAIN_FILL:
                cell.fill = make_fill(DOMAIN_FILL[domain])

        # Difficulty colour (col H = 8)
        elif col_idx == 8:
            diff = value.lower()
            if diff in DIFF_COLOR:
                cell.font = make_font(color=DIFF_COLOR[diff], name="Arial", size=10, bold=True)

        # In Scope colour (col G = 7)
        elif col_idx == 7:
            if value in SCOPE_FILL:
                cell.fill = make_fill(SCOPE_FILL[value])

# Column widths
for col_idx, width in COL_WIDTHS.items():
    ws1.column_dimensions[get_column_letter(col_idx)].width = width

# Row heights for data rows — Question and Notes columns need height
for row_idx in range(2, len(questions) + 2):
    ws1.row_dimensions[row_idx].height = 40

# Freeze top row
ws1.freeze_panes = "A2"

# Auto-filter
ws1.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

# ============================================================
# SHEET 2 — Summary Dashboard
# ============================================================
ws2 = wb.create_sheet(title="Summary Dashboard")
ws2.sheet_view.showGridLines = False

def set_cell(ws, row, col, value, bold=False, font_color="000000",
             fill_color=None, font_size=11, align="left", wrap=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=font_color, name="Arial", size=font_size)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if fill_color:
        cell.fill = make_fill(fill_color)
    cell.border = make_border()
    return cell

# Title
title_cell = ws2.cell(row=1, column=1,
    value="CloudFuze Manage AI Agent — Eval Question Summary")
title_cell.font = Font(bold=True, color="1E3A5F", name="Arial", size=14)
title_cell.alignment = Alignment(horizontal="left", vertical="center")
ws2.merge_cells("A1:C1")
ws2.row_dimensions[1].height = 30

# Helper for table headers
def table_header_row(ws, row, labels):
    for col_idx, label in enumerate(labels, 1):
        set_cell(ws, row, col_idx, label,
                 bold=True, font_color="FFFFFF", fill_color="1E3A5F",
                 align="center")

# ── Table 1: By Domain (rows 3–12) ───────────────────────────────────────────
table_header_row(ws2, 3, ["Domain", "Count", "% of Total"])

domains = ["licence", "spend", "shadow_it", "contracts",
           "users", "compliance", "multi", "out_of_scope"]
for i, domain in enumerate(domains):
    data_row = 4 + i
    set_cell(ws2, data_row, 1, domain, align="left")
    set_cell(ws2, data_row, 2,
             f"=COUNTIF('All Questions'!C:C,\"{domain}\")",
             align="center")
    set_cell(ws2, data_row, 3,
             f"=IF(B{data_row+8}>0,B{data_row}/B{4+len(domains)},0)",
             align="center")
    # Domain colour in col A
    if domain in DOMAIN_FILL:
        ws2.cell(row=data_row, column=1).fill = make_fill(DOMAIN_FILL[domain])

# Total row (row 12)
total_row_domain = 4 + len(domains)
set_cell(ws2, total_row_domain, 1, "Total", bold=True, align="left")
set_cell(ws2, total_row_domain, 2,
         f"=SUM(B4:B{total_row_domain - 1})", bold=True, align="center")
set_cell(ws2, total_row_domain, 3, "100%", bold=True, align="center")

# Fix % formulas now that we know total row
for i in range(len(domains)):
    data_row = 4 + i
    ws2.cell(row=data_row, column=3).value = (
        f"=IF(B{total_row_domain}>0,B{data_row}/B{total_row_domain},0)"
    )
    ws2.cell(row=data_row, column=3).number_format = "0.0%"

# ── Table 2: By Difficulty (rows 14–19) ──────────────────────────────────────
table_header_row(ws2, 14, ["Difficulty", "Count", "% of Total"])

difficulties = ["easy", "medium", "hard"]
for i, diff in enumerate(difficulties):
    data_row = 15 + i
    cell_a = set_cell(ws2, data_row, 1, diff, align="left")
    if diff in DIFF_COLOR:
        cell_a.font = Font(bold=True, color=DIFF_COLOR[diff], name="Arial", size=11)
    set_cell(ws2, data_row, 2,
             f"=COUNTIF('All Questions'!H:H,\"{diff}\")",
             align="center")
    set_cell(ws2, data_row, 3,
             f"=IF(B19>0,B{data_row}/B19,0)",
             align="center")
    ws2.cell(row=data_row, column=3).number_format = "0.0%"

# Total row 19
set_cell(ws2, 19, 1, "Total", bold=True, align="left")
set_cell(ws2, 19, 2, "=SUM(B15:B17)", bold=True, align="center")
set_cell(ws2, 19, 3, "100%", bold=True, align="center")

# ── Table 3: By Intent (rows 21–29) ──────────────────────────────────────────
table_header_row(ws2, 21, ["Intent", "Count", ""])

intents = ["lookup", "aggregate", "compare", "forecast", "action", "out_of_scope"]
for i, intent in enumerate(intents):
    data_row = 22 + i
    set_cell(ws2, data_row, 1, intent, align="left")
    set_cell(ws2, data_row, 2,
             f"=COUNTIF('All Questions'!D:D,\"{intent}\")",
             align="center")
    set_cell(ws2, data_row, 3, "", align="center")

# ── Table 4: In Scope vs Out (rows 31–35) ────────────────────────────────────
table_header_row(ws2, 31, ["Type", "Count", ""])

set_cell(ws2, 32, 1, "In Scope (TRUE)", align="left")
ws2.cell(row=32, column=1).fill = make_fill("DCFCE7")
set_cell(ws2, 32, 2, "=COUNTIF('All Questions'!G:G,\"TRUE\")", align="center")
set_cell(ws2, 32, 3, "", align="center")

set_cell(ws2, 33, 1, "Out of Scope (FALSE)", align="left")
ws2.cell(row=33, column=1).fill = make_fill("FEE2E2")
set_cell(ws2, 33, 2, "=COUNTIF('All Questions'!G:G,\"FALSE\")", align="center")
set_cell(ws2, 33, 3, "", align="center")

set_cell(ws2, 34, 1, "Total", bold=True, align="left")
set_cell(ws2, 34, 2, "=SUM(B32:B33)", bold=True, align="center")
set_cell(ws2, 34, 3, "", align="center")

# ── Table 5: By Expected Widget (rows 37–46) ─────────────────────────────────
table_header_row(ws2, 37, ["Widget Type", "Count", ""])

widgets = ["metric_cards", "table", "bar_chart", "donut_chart",
           "timeline", "text_block", "action_buttons"]
for i, widget in enumerate(widgets):
    data_row = 38 + i
    set_cell(ws2, data_row, 1, widget, align="left")
    set_cell(ws2, data_row, 2,
             f"=COUNTIF('All Questions'!E:E,\"{widget}\")",
             align="center")
    set_cell(ws2, data_row, 3, "", align="center")

# ── Summary sheet column widths ──────────────────────────────────────────────
ws2.column_dimensions["A"].width = 25
ws2.column_dimensions["B"].width = 12
ws2.column_dimensions["C"].width = 14

# ── Save ─────────────────────────────────────────────────────────────────────
wb.save(EXCEL_PATH)
print(f"Saved: {EXCEL_PATH}")
